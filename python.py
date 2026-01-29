import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from itertools import product
from io import BytesIO

# PDF (pip install reportlab)
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

# ============================================================
# CONFIG
# ============================================================
EXCEL_PATH = "Chinese Characters.xlsx"

FIRST_CHAR = {
    "char": "洪",
    "pinyin": "hóng",
    "element": "木",
    "strokes": 10
}

# Destiny meaning uses TOTAL strokes WITHOUT +1
DESTINY_MEANINGS = {
    16: {
        "zh": "（吉）能夠克己助人而敦厚雅量，安富會榮而福壽雙全。女性則能益夫興家而子孫榮昌，並且賢淑而理家有方。",
        "en": "(Auspicious) Self-disciplined and generous, with refined magnanimity. Enjoys stability, prosperity, and longevity. "
              "For women: supports spouse, prospers the home, and manages family affairs wisely; descendants flourish."
    },
    25: {
        "zh": "（吉）能得天時與地利，但難以得人和，學識豐富而精神活潑，有領導的能力，若能善用人才則大可成功。至於女性則很有才氣，並且溫和賢淑而有感情。",
        "en": "(Auspicious) Blessed with favorable timing and conditions, though harmony with people needs effort. "
              "Knowledgeable and lively with leadership ability; great success comes by using talent well. "
              "For women: gifted, gentle, and affectionate."
    },
    31: {
        "zh": "（吉）能夠腳踏實地而智勇雙全，性情溫和而寬宏大量，貴人明現。",
        "en": "(Auspicious) Grounded and steadfast, possessing both wisdom and courage. Gentle and broad-minded, "
              "with benefactors appearing at key times."
    },
    41: {
        "zh": "（吉）有才能也有理智，前程似錦，有官運與財運。",
        "en": "(Auspicious) Talented and sensible, with a bright future. Strong career/authority luck and wealth fortune are indicated."
    },
    45: {
        "zh": "（吉）做事一帆風順，智勇雙全必能有所成就。女性則不要虛榮則吉，可助夫益子而使家業興隆。",
        "en": "(Auspicious) Endeavors proceed smoothly; with wisdom and courage, achievement is assured. "
              "For women: avoid vanity for better fortune; supports spouse and children, prospering the family estate."
    }
}

PATTERN_MEANINGS = {
    "木木木": {
        "en": "The foundation is stable. Wishes can largely be fulfilled; prosperity and longevity follow.",
        "zh": "基礎安定，所求之事頗能如願，家業興隆而身心健全，保養得宜則能長壽。"
    },
    "木木土": {
        "en": "A steady temperament and solid fortune; health, happiness, and longevity. Be forgiving with others.",
        "zh": "性情穩健，境遇堅固，身心健康而幸福長壽；與人相處宜寬恕。"
    },
    "木火水": {
        "en": "Supported by both superiors and subordinates, enabling steady growth and advancement; happiness and longevity.",
        "zh": "能得到上下的幫助而發展，幸福長壽。"
    },
    "木火土": {
        "en": "Guided by elders; warm, sincere, and popular—thus happiness and longevity.",
        "zh": "受長輩引進而發展成功，對人熱情，對下更親切而有人緣，因此長壽幸福。"
    },
    "木土火": {
        "en": "Strong interpersonal harmony; a sound foundation supports successful growth and advancement.",
        "zh": "有人緣，基礎運健全，能成功發展。"
    }
}

# Requested stroke tuples ONLY (second/third do NOT need to match any element in Excel)
REQUESTED_COMBOS = {
    "木木木": [(11, 10), (1, 20), (11, 20), (21, 10)],
    "木木土": [(21, 14), (1, 5), (11, 24)],
    "木火土": [(3, 12)],
}

# Destiny totals filter (NO +1)
PATTERN_TOTAL_FILTERS = {
    "木木木": {31, 41},
    "木木土": {16, 45},
    "木火土": {25}
}

# ============================================================
# UI COLORS FOR FIVE ELEMENTS
# ============================================================
ELEMENT_COLORS = {
    "木": "#2E7D32",  # Green
    "火": "#C62828",  # Red
    "土": "#8D6E63",  # Brown
    "金": "#9E9E9E",  # Gray
    "水": "#1565C0",  # Blue
}

def five_grid_tooltip(key: str, lang: str) -> str:
    tips = {
        "天格": {
            "en": "Heaven Grid: family background, ancestors, early influence",
            "zh": "天格：祖先、家族背景、早年影響"
        },
        "人格": {
            "en": "Personality Grid: core character, talent, life direction",
            "zh": "人格：主運、性格、才能與人生方向"
        },
        "地格": {
            "en": "Earth Grid: early life, relationships, foundation",
            "zh": "地格：前運、人際關係與基礎"
        },
        "總格": {
            "en": "Total Grid: overall destiny (calculated WITHOUT +1)",
            "zh": "總格：一生命運（不加1）"
        }
    }

    if lang == "English":
        return tips[key]["en"]
    if lang == "Chinese":
        return tips[key]["zh"]
    return f"{tips[key]['en']} ｜ {tips[key]['zh']}"

def element_badge(element: str) -> str:
    color = ELEMENT_COLORS.get(element, "#333333")
    return f"<span style='color:{color}; font-weight:700;'>{element}</span>"

# ============================================================
# STROKE → ELEMENT (last digit rule)
# ============================================================
def stroke_to_element(strokes: int) -> str:
    last = strokes % 10
    if last in (1, 2):
        return "木"
    if last in (3, 4):
        return "火"
    if last in (5, 6):
        return "土"
    if last in (7, 8):
        return "金"
    return "水"

# ============================================================
# 五格 (天格/人格/地格/總格)
# ============================================================
def compute_five_grids(first: int, second: int, third: int):
    tian = first + 1
    ren = first + second
    di = second + third
    zong = first + second + third 
    return {
        "天格": (tian, stroke_to_element(tian)),
        "人格": (ren, stroke_to_element(ren)),
        "地格": (di, stroke_to_element(di)),
        "總格": (zong, stroke_to_element(zong)),
    }

# ============================================================
# Pattern elements (your +1 rule)
# ============================================================
def compute_pattern_elements(first: int, second: int, third: int):
    A = first + 1
    B = first + second
    C = second + third
    return {
        "calc_text": f"{first}+1={A}({stroke_to_element(A)}) · {first}+{second}={B}({stroke_to_element(B)}) · {second}+{third}={C}({stroke_to_element(C)})",
        "elements": stroke_to_element(A) + stroke_to_element(B) + stroke_to_element(C),
        "A": A, "B": B, "C": C
    }

# ============================================================
# LOAD DATABASE
# row[0]=char row[1]=pinyin row[2]=strokes row[3]=element row[4]=EN row[5]=ZH
# ============================================================
@st.cache_data(show_spinner=False)
def load_db(excel_path: str):
    wb = load_workbook(excel_path)
    ws = wb.active

    db = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        try:
            char = row[0]
            pinyin = row[1]
            strokes = int(row[2])
            element = row[3]
            meaning_en = row[4] if len(row) > 4 else ""
            meaning_zh = row[5] if len(row) > 5 else ""
            if char and pinyin and strokes and element is not None:
                db.append({
                    "char": char,
                    "pinyin": pinyin,
                    "strokes": strokes,
                    "element": element,
                    "meaning_en": meaning_en or "",
                    "meaning_zh": meaning_zh or ""
                })
        except:
            continue

    # Index ONLY by strokes (ignore element requirement for 2nd/3rd chars)
    by_strokes = {}
    by_char = {}
    for c in db:
        by_strokes.setdefault(c["strokes"], []).append(c)
        by_char[c["char"]] = c

    return db, by_strokes, by_char

# ============================================================
# FILTERS
# ============================================================
def allowed_destiny_total(pattern_key: str, destiny_total: int) -> bool:
    allowed = PATTERN_TOTAL_FILTERS.get(pattern_key)
    return True if not allowed else destiny_total in allowed

# ============================================================
# BUILD RESULT ROW
# ============================================================
def make_row(requested_pattern_key, second, third, by_char):
    first = FIRST_CHAR["strokes"]
    s2 = second["strokes"]
    s3 = third["strokes"]

    destiny_total = first + s2 + s3  # NO +1
    if not allowed_destiny_total(requested_pattern_key, destiny_total):
        return None

    pat = compute_pattern_elements(first, s2, s3)
    computed_pattern = pat["elements"]

    # STRICT pattern match
    if computed_pattern != requested_pattern_key:
        return None

    name = FIRST_CHAR["char"] + second["char"] + third["char"]
    pinyin = f"{FIRST_CHAR['pinyin']} {second['pinyin']} {third['pinyin']}"

    five_grids = compute_five_grids(first, s2, s3)

    first_info = by_char.get(FIRST_CHAR["char"], {
        "char": FIRST_CHAR["char"],
        "pinyin": FIRST_CHAR["pinyin"],
        "strokes": FIRST_CHAR["strokes"],
        "element": FIRST_CHAR["element"],
        "meaning_en": "",
        "meaning_zh": ""
    })
    char_details = [first_info, second, third]

    return {
        "PatternRequested": requested_pattern_key,
        "PatternComputed": computed_pattern,
        "Name": name,
        "Pinyin": pinyin,

        "FiveGrids": five_grids,
        "PatternCalc": pat["calc_text"],

        "DestinyTotal": destiny_total,
        "DestinyElement": stroke_to_element(destiny_total),
        "DestinyMeaning_EN": DESTINY_MEANINGS.get(destiny_total, {}).get("en", "Not defined."),
        "DestinyMeaning_ZH": DESTINY_MEANINGS.get(destiny_total, {}).get("zh", "（未定義）"),

        "PatternMeaning_EN": PATTERN_MEANINGS.get(computed_pattern, {}).get("en", ""),
        "PatternMeaning_ZH": PATTERN_MEANINGS.get(computed_pattern, {}).get("zh", ""),

        "CharDetails": char_details,
    }

# ============================================================
# FAVORITES + PDF
# ============================================================
def ensure_state():
    if "favorites" not in st.session_state:
        st.session_state.favorites = []
ensure_state()

def add_favorite(row_dict):
    # de-duplicate by Name
    for f in st.session_state.favorites:
        if f.get("Name") == row_dict.get("Name"):
            return False
    st.session_state.favorites.append(row_dict)
    return True

def remove_favorite(name: str):
    st.session_state.favorites = [f for f in st.session_state.favorites if f.get("Name") != name]

def clear_favorites():
    st.session_state.favorites = []

def _wrap_text(canvas_obj, text, x, y, max_width, line_height=14, font_name="Helvetica", font_size=10):
    """
    Simple word-wrapping for reportlab.
    """
    canvas_obj.setFont(font_name, font_size)
    words = (text or "").split()
    line = ""
    lines = []
    for w in words:
        test = (line + " " + w).strip()
        if canvas_obj.stringWidth(test, font_name, font_size) <= max_width:
            line = test
        else:
            if line:
                lines.append(line)
            line = w
    if line:
        lines.append(line)

    for ln in lines:
        canvas_obj.drawString(x, y, ln)
        y -= line_height
    return y

def generate_pdf(favorites, lang_mode="Both"):
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 40

    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "Chinese Name Analysis Report")
    y -= 18
    c.setFont("Helvetica", 10)
    c.drawString(40, y, "Favorites comparison export (五格・五行組合・總格數理)")
    y -= 24

    for idx, f in enumerate(favorites, start=1):
        if y < 140:
            c.showPage()
            y = height - 40

        c.setFont("Helvetica-Bold", 12)
        c.drawString(40, y, f"{idx}. {f['Name']}  ({f['Pinyin']})")
        y -= 16

        fg = f["FiveGrids"]
        c.setFont("Helvetica", 10)
        c.drawString(40, y, f"Pattern (組合): {f['PatternComputed']}   |   Total (總格): {f['DestinyTotal']} ({f['DestinyElement']})")
        y -= 14
        c.drawString(
            40,
            y,
            f"Five Grids 五格: 天格 {fg['天格'][0]}({fg['天格'][1]}) · 人格 {fg['人格'][0]}({fg['人格'][1]}) · 地格 {fg['地格'][0]}({fg['地格'][1]}) · 總格 {fg['總格'][0]}({fg['總格'][1]})"
        )
        y -= 16

        c.setFont("Helvetica-Oblique", 10)
        c.drawString(40, y, "Pattern calculation (+1 rule):")
        y -= 14
        c.setFont("Helvetica", 10)
        y = _wrap_text(c, f.get("PatternCalc", ""), 50, y, max_width=width - 90, line_height=13)

        y -= 6
        c.setFont("Helvetica-Oblique", 10)
        c.drawString(40, y, "Meanings:")
        y -= 14
        c.setFont("Helvetica", 10)

        if lang_mode in ("English", "Both"):
            y = _wrap_text(c, "Pattern (EN): " + (f.get("PatternMeaning_EN") or "—"), 50, y, width - 90)
            y = _wrap_text(c, "Destiny (EN): " + (f.get("DestinyMeaning_EN") or "—"), 50, y, width - 90)
            y -= 4
        if lang_mode in ("Chinese", "Both"):
            y = _wrap_text(c, "組合(中): " + (f.get("PatternMeaning_ZH") or "—"), 50, y, width - 90)
            y = _wrap_text(c, "數理(中): " + (f.get("DestinyMeaning_ZH") or "—"), 50, y, width - 90)
            y -= 4

        # Character details
        c.setFont("Helvetica-Oblique", 10)
        c.drawString(40, y, "Characters:")
        y -= 14
        c.setFont("Helvetica", 10)

        for ch in f.get("CharDetails", []):
            if y < 120:
                c.showPage()
                y = height - 40
                c.setFont("Helvetica", 10)

            line = f"{ch.get('char','')} ({ch.get('pinyin','')}), {ch.get('strokes','')} strokes, element {ch.get('element','')}"
            c.drawString(50, y, line)
            y -= 12
            if lang_mode in ("English", "Both"):
                y = _wrap_text(c, "EN: " + (ch.get("meaning_en") or "—"), 60, y, width - 100, line_height=12)
            if lang_mode in ("Chinese", "Both"):
                y = _wrap_text(c, "中: " + (ch.get("meaning_zh") or "—"), 60, y, width - 100, line_height=12)
            y -= 6

        # divider
        c.line(40, y, width - 40, y)
        y -= 16

    c.save()
    buffer.seek(0)
    return buffer

# ============================================================
# STREAMLIT UI
# ============================================================
st.set_page_config(page_title="（洪）Professional Name Generator", layout="wide")
st.title("🔮（洪）Professional Chinese Name Generator")
st.caption("✅ 第二/第三字只依筆畫配對（不需符合Excel五行）｜✅ 組合五行依 +1 規則｜✅ 總格數理不加 +1")

guide_lang = st.sidebar.radio(
    "Guide Language | 說明語言",
    ["English", "Chinese", "Both"],
    index=0
)

# Guide / help (bilingual)
with st.expander("📘 How to Read This Name Analysis | 使用說明（必讀）", expanded=True):

    if guide_lang in ("English", "Both"):
        st.markdown("## 🌟 English Guide")
        st.markdown("""
**This tool generates Chinese names based on traditional name numerology (姓名學) and Five-Element theory (五行).**

### 1️⃣ Five Grids (五格)
- **Heaven Grid (天格)** – family background and ancestral influence  
- **Personality Grid (人格)** – core personality, talent, and life direction  
- **Earth Grid (地格)** – early life, relationships, and foundation  
- **Total Grid (總格)** – overall destiny (**NO +1**)

---

### 2️⃣ Five-Element Pattern (五行組合)
Patterns such as **木木木 / 木木土 / 木火土** are calculated using:
- Surname strokes + 1  
- Surname + first given name  
- First + second given name  

---

### 3️⃣ Destiny Meaning (數理)
Destiny meaning is determined **only by total strokes**, without adding +1.

---

### 4️⃣ Character Meanings
Each character displays pinyin, strokes, element, and English/Chinese meanings from the database.
""")

    if guide_lang in ("Chinese", "Both"):
        st.markdown("## 🌟 中文說明")
        st.markdown("""
**本系統依據傳統姓名學與五行數理設計，適合一般使用者閱讀與理解。**

### 1️⃣ 五格說明
- **天格**：祖先、家族背景  
- **人格**：主運、性格與才能  
- **地格**：前運、人際關係  
- **總格**：一生命運（**總格不加1**）

---

### 2️⃣ 五行組合
如「木木木、木木土、木火土」等組合，計算方式如下：
- 姓氏筆畫 + 1  
- 姓氏 + 名字第一字  
- 名字第一字 + 第二字  

---

### 3️⃣ 數理含義
數理僅以三字總筆畫判斷，不加1。

---

### 4️⃣ 單字含義
展開卡片即可查看每個字的拼音、筆畫、五行，以及中英文含義。
""")

db, by_strokes, by_char = load_db(EXCEL_PATH)

# Sidebar controls
st.sidebar.header("Controls")
lang = st.sidebar.radio("Meaning Language", ["English", "Chinese", "Both"], 0)
show_destiny = st.sidebar.toggle("Show destiny meaning (總格數理)", value=True)
limit = st.sidebar.slider("Max cards to show", 10, 5000, 500, step=20)
search = st.sidebar.text_input("Search (Name / Pinyin)", "")

selected_patterns = st.sidebar.multiselect(
    "Select patterns",
    options=list(REQUESTED_COMBOS.keys()),
    default=list(REQUESTED_COMBOS.keys())
)

# Generate
rows = []
for pattern_key in selected_patterns:
    for s2, s3 in REQUESTED_COMBOS.get(pattern_key, []):
        seconds = by_strokes.get(s2, [])
        thirds = by_strokes.get(s3, [])
        if not seconds or not thirds:
            continue
        for second, third in product(seconds, thirds):
            r = make_row(pattern_key, second, third, by_char)
            if r:
                rows.append(r)

df = pd.DataFrame(rows)
if df.empty:
    st.warning("No results found. Check Excel strokes availability, requested tuples, pattern filters, or destiny total filters.")
    st.stop()

# Search
if search.strip():
    q = search.strip().lower()
    df = df[
        df["Name"].astype(str).str.lower().str.contains(q) |
        df["Pinyin"].astype(str).str.lower().str.contains(q)
    ]

# Summary
c1, c2, c3 = st.columns([1.2, 1, 1])
c1.metric("Results", f"{len(df)}")
c2.metric("Patterns", f"{df['PatternComputed'].nunique()}")
c3.metric("Destiny Totals", f"{df['DestinyTotal'].nunique()}")

st.divider()

# Export table
with st.expander("📋 Table view / Export"):
    base_cols = ["PatternComputed", "Name", "Pinyin", "DestinyTotal", "DestinyElement", "PatternCalc"]
    extra_cols = []
    if lang in ("English", "Both"):
        extra_cols += ["PatternMeaning_EN"]
    if lang in ("Chinese", "Both"):
        extra_cols += ["PatternMeaning_ZH"]
    if show_destiny:
        if lang == "English":
            extra_cols += ["DestinyMeaning_EN"]
        elif lang == "Chinese":
            extra_cols += ["DestinyMeaning_ZH"]
        else:
            extra_cols += ["DestinyMeaning_EN", "DestinyMeaning_ZH"]

    show_cols = base_cols + extra_cols
    st.dataframe(df[show_cols], height=360)
    csv = df[show_cols].to_csv(index=False).encode("utf-8-sig")
    st.download_button("Download CSV", data=csv, file_name="name_results.csv", mime="text/csv")

# ============================================================
# NAME CARDS + FAVORITES
# ============================================================
st.subheader("✨ Name Cards")
st.caption("Expand each card to see 五格, 五行組合計算, 總格數理, and each character meaning. Save names to Favorites for comparison and PDF export.")

for r in df.head(limit).to_dict(orient="records"):
    if lang == "English":
        title = (
            f"{r['Name']} · {r['Pinyin']} · "
            f"Total {r['DestinyTotal']} · Pattern {r['PatternComputed']}"
        )
    elif lang == "Chinese":
        title = (
            f"{r['Name']} · {r['Pinyin']} · "
            f"總格 {r['DestinyTotal']} · 組合 {r['PatternComputed']}"
        )
    else:  # Both
        title = (
            f"{r['Name']} · {r['Pinyin']} · "
            f"Total/總格 {r['DestinyTotal']} · Pattern/組合 {r['PatternComputed']}"
        )

    with st.expander(title):

        # Save favorite
        colA, colB = st.columns([1, 5])
        with colA:
            if st.button("⭐ Save", key=f"save_{r['Name']}"):
                ok = add_favorite(r)
                if ok:
                    st.success("Saved to favorites!")
                else:
                    st.info("Already in favorites.")
        with colB:
            st.write("")

        # 五格 / Five Grids header + explanation (language-aware)
        if lang == "English":
            st.markdown("### 🧭 Five Grids (Heaven · Personality · Earth · Total)")

        elif lang == "Chinese":
            st.markdown("### 🧭 五格（天格・人格・地格・總格）")
            st.caption(
                "天格：姓+1 ｜ "
                "人格：姓+名1 ｜ "
                "地格：名1+名2 ｜ "
                "總格：三字總和（總格不加1）"
            )

        else:  # Both
            st.markdown("### 🧭 Five Grids 五格（Heaven・Personality・Earth・Total）")
            st.caption(
                "Heaven Grid 天格：surname + 1 ｜ "
                "Personality Grid 人格：surname + first given name ｜ "
                "Earth Grid 地格：first + second given name ｜ "
                "Total Grid 總格：sum of all three characters (NO +1 / 不加1)"
            )

        fg = r["FiveGrids"]
        cols = st.columns(4)

        for i, key in enumerate(["天格", "人格", "地格", "總格"]):
            strokes, elem = fg[key]

            # Language-aware label
            if lang == "English":
                label_map = {
                    "天格": "Heaven Grid",
                    "人格": "Personality Grid",
                    "地格": "Earth Grid",
                    "總格": "Total Grid",
                }
                label = label_map[key]
            elif lang == "Chinese":
                label = key
            else:  # Both
                label_map = {
                    "天格": "Heaven Grid 天格",
                    "人格": "Personality Grid 人格",
                    "地格": "Earth Grid 地格",
                    "總格": "Total Grid 總格",
                }
                label = label_map[key]

            # Metric with tooltip
            cols[i].metric(
                label=label,
                value=str(strokes),
                delta=elem,
                help=five_grid_tooltip(key, lang)
            )

        st.divider()

        left, right = st.columns([1.05, 1.35])
        with left:
            st.markdown("#### 🔢 Calculations")
            st.write(f"**Pattern calc (+1 rule):** {r['PatternCalc']}")
            st.write(f"**Destiny total (no +1):** {r['DestinyTotal']} → 五行: **{r['DestinyElement']}**")

        with right:
            st.markdown("#### 📖 Meanings")

            # Pattern meaning
            if lang == "English":
                st.markdown("**Pattern Meaning (EN)**")
                st.write(r.get("PatternMeaning_EN", "") or "—")
            elif lang == "Chinese":
                st.markdown("**組合含義（中文）**")
                st.write(r.get("PatternMeaning_ZH", "") or "—")
            else:
                st.markdown("**Pattern Meaning (EN)**")
                st.write(r.get("PatternMeaning_EN", "") or "—")
                st.markdown("**組合含義（中文）**")
                st.write(r.get("PatternMeaning_ZH", "") or "—")

            # Destiny meaning
            if show_destiny:
                st.markdown(f"**Destiny Meaning 總格數理（{r['DestinyTotal']}）**")
                if lang == "English":
                    st.success(r.get("DestinyMeaning_EN", "Not defined."))
                elif lang == "Chinese":
                    st.success(r.get("DestinyMeaning_ZH", "（未定義）"))
                else:
                    st.success(r.get("DestinyMeaning_EN", "Not defined."))
                    st.info(r.get("DestinyMeaning_ZH", "（未定義）"))

        st.divider()
        st.markdown("### 🔤 Character Details（每個字：拼音・筆畫・五行・含義）")

        for ch in r["CharDetails"]:
            ch_char = ch.get("char", "")
            ch_pinyin = ch.get("pinyin", "")
            ch_strokes = ch.get("strokes", "")
            ch_elem = ch.get("element", "")
            ch_en = ch.get("meaning_en", "") or "—"
            ch_zh = ch.get("meaning_zh", "") or "—"

            st.markdown(
                    f"**{ch_char}** · *{ch_pinyin}* · {ch_strokes} strokes · "
                    f"Element: {element_badge(ch_elem)}",
                    unsafe_allow_html=True
                )

            if lang == "English":
                st.write(f"English: {ch_en}")
            elif lang == "Chinese":
                st.write(f"中文: {ch_zh}")
            else:
                st.write(f"English: {ch_en}")
                st.write(f"中文: {ch_zh}")
            st.write("")

# ============================================================
# FAVORITES PANEL + COMPARE + PDF EXPORT
# ============================================================
st.divider()
st.subheader("⭐ Favorites (Save & Compare)")

if not st.session_state.favorites:
    st.info("No favorite names yet. Save names from the cards above.\n\n尚未收藏任何名字，請在上方卡片按 ⭐ Save。")
else:
    fav_cols = st.columns([3, 1, 1])

    with fav_cols[0]:
        st.write(f"Saved favorites: **{len(st.session_state.favorites)}**")

    with fav_cols[1]:
        if st.button("🗑 Clear", key="clear_favs"):
            clear_favorites()
            st.rerun()

    with fav_cols[2]:
        pdf_lang = st.selectbox("PDF Language", ["English", "Chinese", "Both"], index=2)

    fav_df = pd.DataFrame([
        {
            "Name": f["Name"],
            "Pinyin": f["Pinyin"],
            "Pattern": f["PatternComputed"],
            "Total Strokes (總格)": f["DestinyTotal"],
            "Element": f["DestinyElement"],
            "天格": f["FiveGrids"]["天格"][0],
            "人格": f["FiveGrids"]["人格"][0],
            "地格": f["FiveGrids"]["地格"][0],
            "總格": f["FiveGrids"]["總格"][0],
        }
        for f in st.session_state.favorites
    ])
    st.dataframe(fav_df)

    # Remove individual favorites
    st.markdown("#### Remove an item | 刪除單項")
    rm_name = st.selectbox("Select name to remove", [f["Name"] for f in st.session_state.favorites])
    if st.button("Remove selected", key="rm_btn"):
        remove_favorite(rm_name)
        st.rerun()

    # PDF export
    st.markdown("#### 📄 Export PDF | 輸出PDF")
    pdf_data = generate_pdf(st.session_state.favorites, lang_mode=pdf_lang)
    st.download_button(
        "📄 Download Favorites PDF",
        data=pdf_data,
        file_name="Name_Analysis_Report.pdf",
        mime="application/pdf"
    )
