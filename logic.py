import streamlit as st
from openpyxl import load_workbook
from itertools import product
from typing import Dict, List, Tuple, Any, Optional

from rules.zodiac_rules import check_zodiac_tokens
from config import (
    FIRST_CHAR, DESTINY_MEANINGS, PATTERN_MEANINGS,
    REQUESTED_COMBOS, PATTERN_TOTAL_FILTERS
)

# ============================================================
# STROKE → ELEMENT (last digit rule)
# ============================================================
def stroke_to_element(strokes: int) -> str:
    last = strokes % 10
    if last in (1, 2):
        return "木"
    if last in (3, 4):
        return "火"
    if last in (5, 6):
        return "土"
    if last in (7, 8):
        return "金"
    return "水"

# ============================================================
# 五格 (天格/人格/地格/總格)
# ============================================================
def compute_five_grids(first: int, second: int, third: int) -> Dict[str, Tuple[int, str]]:
    tian = first + 1
    ren = first + second
    di = second + third
    zong = first + second + third  # NO +1
    return {
        "天格": (tian, stroke_to_element(tian)),
        "人格": (ren, stroke_to_element(ren)),
        "地格": (di, stroke_to_element(di)),
        "總格": (zong, stroke_to_element(zong)),
    }

# ============================================================
# Pattern elements (+1 rule)
# ============================================================
def compute_pattern_elements(first: int, second: int, third: int) -> Dict[str, Any]:
    A = first + 1
    B = first + second
    C = second + third
    return {
        "calc_text": (
            f"{first}+1={A}({stroke_to_element(A)}) · "
            f"{first}+{second}={B}({stroke_to_element(B)}) · "
            f"{second}+{third}={C}({stroke_to_element(C)})"
        ),
        "elements": stroke_to_element(A) + stroke_to_element(B) + stroke_to_element(C),
        "A": A, "B": B, "C": C
    }

# ============================================================
# FILTERS
# ============================================================
def allowed_destiny_total(pattern_key: str, destiny_total: int) -> bool:
    allowed = PATTERN_TOTAL_FILTERS.get(pattern_key)
    return True if not allowed else destiny_total in allowed

# ============================================================
# LOAD DB
# Excel mapping:
# row[0]=char row[1]=pinyin row[2]=strokes row[3]=element
# row[4]=ZODIAC CELL (single source for ALL zodiac filtering)
# row[5]=meaning_en row[6]=meaning_zh
# ============================================================
def load_db_raw(excel_path: str):
    wb = load_workbook(excel_path)
    ws = wb.active

    db = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        try:
            char = row[0]
            pinyin = row[1]
            strokes = int(row[2])
            element = row[3]

            zodiac_cell = row[4] if len(row) > 4 else ""   # ✅ ONLY row[4]
            meaning_en = row[5] if len(row) > 5 else ""    # ✅ shifted
            meaning_zh = row[6] if len(row) > 6 else ""    # ✅ shifted

            if char and pinyin and strokes and element is not None:
                db.append({
                    "char": char,
                    "pinyin": pinyin,
                    "strokes": strokes,
                    "element": element,
                    "zodiac_cell": zodiac_cell or "",
                    "meaning_en": meaning_en or "",
                    "meaning_zh": meaning_zh or "",
                })
        except:
            continue

    by_strokes = {}
    by_char = {}
    for c in db:
        by_strokes.setdefault(c["strokes"], []).append(c)
        by_char[c["char"]] = c

    return db, by_strokes, by_char

# ============================================================
# BUILD RESULT ROW
# zodiac_name chooses rule set; we still read the same row[4] cell.
# filter applies ONLY on 2nd + 3rd chars.
# ============================================================
def make_row(
    requested_pattern_key: str,
    second: dict,
    third: dict,
    by_char: dict,
    zodiac_name: str = "None",
    zodiac_filter_mode: str = "OFF",  # OFF | EXCLUDE_XIONG | REQUIRE_JI
) -> Optional[dict]:
    first = FIRST_CHAR["strokes"]
    s2 = second["strokes"]
    s3 = third["strokes"]

    destiny_total = first + s2 + s3  # NO +1
    if not allowed_destiny_total(requested_pattern_key, destiny_total):
        return None

    pat = compute_pattern_elements(first, s2, s3)
    computed_pattern = pat["elements"]
    if computed_pattern != requested_pattern_key:
        return None

    first_info = by_char.get(FIRST_CHAR["char"], {
        "char": FIRST_CHAR["char"],
        "pinyin": FIRST_CHAR["pinyin"],
        "strokes": FIRST_CHAR["strokes"],
        "element": FIRST_CHAR["element"],
        "meaning_en": "",
        "meaning_zh": "",
        "zodiac_cell": "",
    })

    char_details = [first_info, second, third]

    # build zodiac checks from row[4]
    zodiac_checks = []
    if zodiac_name != "None":
        for ch in char_details:
            cell_text = ch.get("zodiac_cell", "")
            res = check_zodiac_tokens(zodiac_name, cell_text)  # ✅ rule set chosen by zodiac_name
            zodiac_checks.append({
                "char": ch.get("char", ""),
                "status": res.get("status", "neutral"),  # 吉 / 凶 / neutral
                "matched": res.get("matched", ""),
                "cell": cell_text,
            })
    else:
        zodiac_checks = [
            {"char": ch.get("char", ""), "status": "neutral", "matched": "", "cell": ""}
            for ch in char_details
        ]

    # ✅ apply filter ONLY on 2nd + 3rd
    if zodiac_name != "None" and zodiac_filter_mode != "OFF":
        s2_status = zodiac_checks[1]["status"]
        s3_status = zodiac_checks[2]["status"]

        if zodiac_filter_mode == "REQUIRE_JI":
            if s2_status != "吉" or s3_status != "吉":
                return None

        elif zodiac_filter_mode == "EXCLUDE_XIONG":
            if s2_status == "凶" or s3_status == "凶":
                return None

    name = FIRST_CHAR["char"] + second["char"] + third["char"]
    pinyin = f"{FIRST_CHAR['pinyin']} {second['pinyin']} {third['pinyin']}"
    five_grids = compute_five_grids(first, s2, s3)

    return {
        "PatternRequested": requested_pattern_key,
        "PatternComputed": computed_pattern,
        "Name": name,
        "Pinyin": pinyin,
        "FiveGrids": five_grids,
        "PatternCalc": pat["calc_text"],
        "DestinyTotal": destiny_total,
        "DestinyElement": stroke_to_element(destiny_total),
        "DestinyMeaning_EN": DESTINY_MEANINGS.get(destiny_total, {}).get("en", "Not defined."),
        "DestinyMeaning_ZH": DESTINY_MEANINGS.get(destiny_total, {}).get("zh", "（未定義）"),
        "PatternMeaning_EN": PATTERN_MEANINGS.get(computed_pattern, {}).get("en", ""),
        "PatternMeaning_ZH": PATTERN_MEANINGS.get(computed_pattern, {}).get("zh", ""),
        "CharDetails": char_details,
        "ZodiacCheck": {
            "zodiac": zodiac_name,
            "checks": zodiac_checks,
            "filter_mode": zodiac_filter_mode,
        },
    }

# ============================================================
# GENERATE ROWS
# ============================================================
def generate_rows(
    by_strokes: dict,
    by_char: dict,
    selected_patterns: List[str],
    zodiac_name: str = "None",
    zodiac_filter_mode: str = "OFF",
    max_rows: int | None = None, 
) -> List[dict]:
    rows = []
    for pattern_key in selected_patterns:
        for s2, s3 in REQUESTED_COMBOS.get(pattern_key, []):
            seconds = by_strokes.get(s2, [])
            thirds = by_strokes.get(s3, [])
            if not seconds or not thirds:
                continue
            for second, third in product(seconds, thirds):
                r = make_row(
                    pattern_key, second, third, by_char,
                    zodiac_name=zodiac_name,
                    zodiac_filter_mode=zodiac_filter_mode,
                )
                if r:
                    rows.append(r)
                    if max_rows is not None and len(rows) >= max_rows:
                        return rows
    return rows

@st.cache_data(show_spinner=False)
def generate_rows_cached(by_strokes, by_char, selected_patterns, zodiac_name, zodiac_filter_mode, max_rows):
    # by_strokes/by_char are dicts (pickleable) → cache works
    return generate_rows(
        by_strokes, by_char, selected_patterns,
        zodiac_name=zodiac_name,
        zodiac_filter_mode=zodiac_filter_mode,
        max_rows=max_rows,
    )
